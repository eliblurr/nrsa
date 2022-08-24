from datetime import date, datetime
from io import StringIO
from django.db import IntegrityError

#from .data_upload_models import models as data

import nsra.base.models as data
import logging
import csv

from .data_upload_models import *

def format_1(csv_str):
	# logging.info(f"CSV: {csv_str}")
	buffer = StringIO(csv_str)
	csv_reader = csv.DictReader(buffer)

	for row in csv_reader:
		try:
			reportedCase = data.ReportedCase(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				fatal=int(row['fatal'] or 0),
				serious=int(row['serious'] or 0),
				minor=int(row['minor'] or 0))
			reportedCase.save()

			commercial = data.Commercial(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				bus=int(row['com_bus'] or 0),
				minor_bus=int(row['con_truc'] or 0),
				taxi=int(row['com_taxi'] or 0),
				other=int(row['con_others'] or 0),
				injured=int(row['vic_injured'] or 0),
				killed=int(row['vic_killed'] or 0))
			commercial.save()

			private = data.Private(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				minibus=int(row['priv_mini_bus'] or 0),
				salon=int(row['priv_saloon'] or 0),
				suv=int(row['priv_suv/4x4'] or 0),
				govt=int(row['priv_gov'] or 0),
				truck=int(row['priv_truck'] or 0),
				injured=int(row['priv_vic_injured'] or 0),
				killed=int(row['priv_vic_killed'] or 0))
			private.save()

			cycle = data.Cycle(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				m_cycle=int(row['motocycles'] or 0),
				bicycles=int(row['bycicles'] or 0),
				handcart=int(row['handcart'] or 0),
				tricycle=int(row['trycycles'] or 0),
				injured=int(row['moto_vic_injured'] or 0),
				killed=int(row['moto_vic_killed']))
			cycle.save()

			pedestrian = data.Pedestrian(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				injured=int(row['pedestrain_vic_injured'] or 0),
				killed=int(row['pedestrain_vic_killed'] or 0))
			pedestrian.save()

			injured = data.Injured(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				injured=int(row['persons_innjured'] or 0))
			injured.save()

			totalKilled = data.TotalKilled(
				date=date(int(row["year"]), datetime.strptime(row["month"], "%B").month, 1),
				region=row['region'],
				year=row['year'],
				month=row['month'].capitalize(),
				male_over18=int(row['male_above18_killed'] or 0),
				male_under18=int(row['male_below18_killed'] or 0),
				female_over18=int(row['female_above18_killed'] or 0),
				female_under18=int(row['female_below18_killed'] or 0),
				injured=int(row['persons_innjured'] or 0),
				killed=int(row['killed_total'] or 0))
			totalKilled.save()

		# logging.info(f'Current: {row}')

		except Exception as e:
			logging.info(f'Fail: {e} at;\n{row}')
			return f'None'

def format_2(file, table):
	import openpyxl

	workbook = openpyxl.load_workbook(file)
	sheet = workbook[workbook.sheetnames[0]]

	def Road_Env(model, worksheet, start, col):  # for tables with Road Environment Data
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				urban=worksheet.cell(row=rows, column=col + 1).value,
				non_urban=worksheet.cell(row=rows, column=col + 2).value,
			)
			item.save()

	def Vehicles(model, worksheet, start, col):  # For Tables with Vehicular Data
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				Ped=worksheet.cell(row=rows, column=col + 1).value,
				Car=worksheet.cell(row=rows, column=col + 2).value,
				HGV_LGV=worksheet.cell(row=rows, column=col + 3).value,
				Bus_and_Bus_Mini=worksheet.cell(row=rows, column=col + 4).value,
				M_cycle_Tricycle_and_Rickshaw=worksheet.cell(row=rows, column=col + 5).value,
				Pick_Up=worksheet.cell(row=rows, column=col + 6).value,
				Cycle=worksheet.cell(row=rows, column=col + 7).value,
				Other=worksheet.cell(row=rows, column=col + 8).value,
			)
			item.save()

	def Age_Groups(model, worksheet, start, col):  # For Tables with Age Group Data
		for rows in range(start, start + 30):
			try:
				item = model(
					year=worksheet.cell(row=rows, column=col).value,
					Ages_0_to_5=worksheet.cell(row=rows, column=col + 1).value,
					Ages_6_to_15=worksheet.cell(row=rows, column=col + 2).value,
					Ages_16_to_25=worksheet.cell(row=rows, column=col + 3).value,
					Ages_28_to_35=worksheet.cell(row=rows, column=col + 4).value,
					Ages_36_to_45=worksheet.cell(row=rows, column=col + 5).value,
					Ages_46_to_55=worksheet.cell(row=rows, column=col + 6).value,
					Ages_58_to_65=worksheet.cell(row=rows, column=col + 7).value,
					Ages_Over_65=worksheet.cell(row=rows, column=col + 8).value,
				)
				item.save()
			except IntegrityError as e:
				logging.info(f'Data Incomplete after row: {rows}')

	def Sex(model, worksheet, start, col):  # For Tables with Sex Data
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				Male=worksheet.cell(row=rows, column=col + 1).value,
				Female=worksheet.cell(row=rows, column=col + 2).value,
			)
			item.save()

	def Table1_1(model, worksheet):  # for Table 1.1 - Changes in National Traffic Fatality Indices
		for rows in range(8, 38):
			item = model(
				year=worksheet.cell(row=rows, column=1).value,
				all_crashes=worksheet.cell(row=rows, column=2).value,
				all_casualties=worksheet.cell(row=rows, column=3).value,
				fatalities=worksheet.cell(row=rows, column=4).value,
				estimated_population=worksheet.cell(row=rows, column=5).value,
				registered_vehicles=worksheet.cell(row=rows, column=6).value,
				fatalities_per_10k_vehicles=worksheet.cell(row=rows, column=7).value,
				fatalities_per_100k_population=worksheet.cell(row=rows, column=8).value,
				fatalities_per_100_casualties=worksheet.cell(row=rows, column=9).value,
				fatalities_per_100_crashes=worksheet.cell(row=rows, column=10).value,
			)
			item.save()

	def Table1_5(model, worksheet, start, col):  # for Table 1.5 - Vehicle Type Involved in Crashes
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				Car=worksheet.cell(row=rows, column=col + 1).value,
				HGV_LGV=worksheet.cell(row=rows, column=col + 2).value,
				Bus_and_Bus_Mini=worksheet.cell(row=rows, column=col + 3).value,
				M_cycle_Tricycle_and_Rickshaw=worksheet.cell(row=rows, column=col + 4).value,
				Pick_Up=worksheet.cell(row=rows, column=col + 5).value,
				Cycle=worksheet.cell(row=rows, column=col + 6).value,
				Other=worksheet.cell(row=rows, column=col + 7).value,
			)
			item.save()

	def Table1_2_1(model, worksheet, start, col):
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				all_crashes=worksheet.cell(row=rows, column=col + 1).value,
				fatal_crashes=worksheet.cell(row=rows, column=col + 3).value,
				injury_crashes=worksheet.cell(row=rows, column=col + 5).value,
				damage_only=worksheet.cell(row=rows, column=col + 7).value
			)
			item.save()

	def Table1_2_2(model, worksheet, start, col):
		for rows in range(start, start + 30):
			item = model(
				year=worksheet.cell(row=rows, column=col).value,
				all_casualties=worksheet.cell(row=rows, column=col + 1).value,
				killed=worksheet.cell(row=rows, column=col + 3).value,
				seriously_injured=worksheet.cell(row=rows, column=col + 5).value,
				slightly_injured=worksheet.cell(row=rows, column=col + 7).value
			)
			item.save()

	if table.lower() == 'table 1.1':
		#the_class = globals()['Changes_in_National_Traffic_Fatality_Indices']
		the_class = Changes_in_National_Traffic_Fatality_Indices
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Table1_1(the_class, sheet)
			logging.info('Table 1.1 Saved!')

	elif table.lower() == 'table 1.3.1':
		#the_class = globals()['Annual_Distribution_of_Fatalities_by_Road_Environment']
		the_class = Annual_Distribution_of_Fatalities_by_Road_Environment
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Road_Env(
				the_class,
				sheet, 5, 1)
			logging.info('Table 1.3.1 Saved!')

	elif table.lower() == 'table 1.3.4':
		the_class = globals()['Annual_Distribution_of_Non_Urban_Fatalities_by_Road_User_Class']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Vehicles(
				the_class,
				sheet, 6, 1
			)
			logging.info('Table 1.3.4 Saved!')

	elif table.lower() == 'table 1.3.3':
		the_class = globals()['Annual_Distribution_of_Urban_Fatalities_by_Road_User_Class']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Vehicles(
				the_class,
				sheet, 6, 1
			)
			logging.info('Table 1.3.3 Saved!')

	elif table.lower() == 'table 1.3.5':
		the_class = globals()['Annual_Distribution_of_Fatalities_by_Age_Group']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Age_Groups(
				the_class,
				sheet, 7, 1
			)
			logging.info('Table 1.3.5 Saved!')

	elif table.lower() == 'table 1.3.6':
		the_class = globals()['Annual_Distribution_of_Fatalities_by_Sex']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Sex(
				the_class,
				sheet, 7, 2
			)
			logging.info('Table 1.3.6 Saved!')

	elif table.lower() == 'table 1.4.1':
		the_class = globals()['Annual_Distribution_of_Casualties_by_Road_User_Class']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Vehicles(
				the_class,
				sheet, 6, 1
			)
			logging.info('Table 1.4.1 Saved!')

	elif table.lower() == 'table 1.4.2':
		the_class = globals()['Annual_Distribution_of_Urban_Casualties_by_Road_User_Class']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Vehicles(
				the_class,
				sheet, 10, 1
			)
			logging.info('Table 1.4.2 Saved!')

	elif table.lower() == 'table 1.4.3':
		the_class = globals()['Annual_Distribution_of_Non_Urban_Casualties_by_Road_User_Class']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Vehicles(
				the_class,
				sheet, 9, 2
			)
			logging.debug('Table 1.4.3 received')

	elif table.lower() == 'table 1.4.4':
		the_class = globals()['Annual_Distribution_of_Casualties_by_Age_Group']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Age_Groups(
				globals()['Annual_Distribution_of_Casualties_by_Age_Group'],
				sheet, 5, 1
			)
			logging.debug('Table 1.4.4 received')

	elif table.lower() == 'table 1.4.5':
		the_class = globals()['Annual_Distribution_of_Casualties_by_Road_Environment']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Road_Env(
				the_class,
				sheet, 9, 1
			)
			logging.debug('Table 1.4.5 received')

	elif table.lower() == 'table 1.4.6':
		the_class = globals()['Annual_Distribution_of_Casualties_by_Sex']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Sex(
				the_class,
				sheet, 5, 1
			)
			logging.debug('Table 1.4.6 received')

	elif table.lower() == 'table 1.5':
		the_class = globals()['Vehicle_Type_Involved_in_Crashes']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Table1_5(
				the_class,
				sheet, 4, 1
			)
			logging.debug('Table 1.5 received')

	elif table.lower() == 'table 1.2.1':
		the_class = globals()['National_Trends_in_Traffic_Crashes']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Table1_2_1(
				the_class,
				sheet, 7, 1
			)
			logging.debug('Table 1.2.1 received')

	elif table.lower() == 'table 1.2.2':
		the_class = globals()['National_Trends_in_Traffic_Casualties']
		counter = the_class.objects.count()
		if counter in range(25, 31):
			logging.info(f'File has already been Uploaded!; items in DB: {counter}')
		else:
			Table1_2_2(
				the_class,
				sheet, 5, 1
			)
			logging.debug('Table 1.2.2 received')

	else:
		logging.warning(f'No specific Extractor found for {table} ')

def main(file, name):
	workbook = openpyxl.load_workbook(file)
	sheets = workbook.sheetnames
	start_row_num = 15
	region_col = 2
	rc_col = 3
	com_col = rc_col + 4
	priv_col = com_col + 8
	cycle_col = priv_col + 8
	ped_col = cycle_col + 8
	killed_col = ped_col + 3

	file_year = int(name.split(' ')[0])

	for month in sheets[:12]:
		the_month = month.capitalize()
		sheet = month.upper()
		the_date = date(file_year, datetime.strptime(the_month, "%B").month, 1)

		# check if data for given year exists in db; store existence as boolean value
		existence = len(list(ReportedCase.objects.filter(year=file_year, month=the_month).values())) != 0
		logging.info(existence)

		if existence:
			logging.info(f'Data already exists for {the_month} {file_year}')
			continue
		else:
			for rows in range(start_row_num, start_row_num + 17):
				rc = data.ReportedCase(
					date=the_date,
					year=file_year,
					month=the_month,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					fatal=workbook.get_sheet_by_name(sheet).cell(row=rows, column=rc_col).value,
					serious=workbook.get_sheet_by_name(sheet).cell(row=rows, column=rc_col + 1).value,
					minor=workbook.get_sheet_by_name(sheet).cell(row=rows, column=rc_col + 2).value
				)
				rc.save()

				com = data.Commercial(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					bus=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col).value,
					minor_bus=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 1).value,
					truck=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 2).value,
					taxi=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 3).value,
					other=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 4).value,
					injured=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 6).value,
					killed=workbook.get_sheet_by_name(sheet).cell(row=rows, column=com_col + 7).value
				)
				com.save()

				private = data.Private(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					minibus=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col).value,
					salon=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 1).value,
					suv=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 2).value,
					truck=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 3).value,
					govt=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 4).value,
					injured=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 6).value,
					killed=workbook.get_sheet_by_name(sheet).cell(row=rows, column=priv_col + 7).value, )
				private.save()

				cycle = data.Cycle(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					m_cycle=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col).value,
					bicycles=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col + 1).value,
					handcart=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col + 2).value,
					tricycle=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col + 3).value,
					injured=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col + 5).value,
					killed=workbook.get_sheet_by_name(sheet).cell(row=rows, column=cycle_col + 6).value
				)
				cycle.save()

				pedestrian = data.Pedestrian(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					injured=workbook.get_sheet_by_name(sheet).cell(row=rows, column=ped_col).value,
					killed=workbook.get_sheet_by_name(sheet).cell(row=rows, column=ped_col + 1).value,
				)
				pedestrian.save()
				# a var to calculate and hold the total/net injuries
				net_injuries = workbook.get_sheet_by_name(
					sheet).cell(row=rows, column=com_col + 6
									).value + workbook.get_sheet_by_name(
					sheet).cell(row=rows, column=priv_col + 6
									).value + workbook.get_sheet_by_name(
					sheet).cell(row=rows, column=cycle_col + 5).value
				injured = Injured(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					injured=net_injuries,
				)
				injured.save()

				# vars in order to calculate the total and then assign it in model
				male_under18 = workbook.get_sheet_by_name(sheet).cell(row=rows, column=killed_col).value
				male_over18 = workbook.get_sheet_by_name(sheet).cell(row=rows, column=killed_col + 1).value
				female_over18 = workbook.get_sheet_by_name(sheet).cell(row=rows, column=killed_col + 2).value
				female_under18 = workbook.get_sheet_by_name(sheet).cell(row=rows, column=killed_col + 3).value
				# the total; a sum of the vars
				kills = male_under18 + male_over18 + female_over18 + female_under18

				totalKilled = data.TotalKilled(
					date=the_date,
					region=workbook.get_sheet_by_name(sheet).cell(row=rows, column=region_col).value,
					year=file_year,
					month=the_month,
					male_under18=male_under18,
					male_over18=male_over18,
					female_over18=female_over18,
					female_under18=female_under18,
					killed=kills,
				)
				totalKilled.save()

	return 0